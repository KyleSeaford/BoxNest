from openpyxl import workbook, load_workbook
import time
import os
import pyautogui
from userQuestions import *
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )


os.system('color 9f')

word = "database"

workbook = load_workbook(f"{word}.xlsx")
sheet = workbook.active

name = r"""
 ____               _   _             _   
| __ )   ___ __  __| \ | |  ___  ___ | |_ 
|  _ \  / _ \\ \/ /|  \| | / _ \/ __|| __|  Organized to Perfection
| |_) || (_) |>  < | |\  ||  __/\__ \| |_   -----------------------
|____/  \___//_/\_\|_| \_| \___||___/ \__|     By: Kyle Seaford                                        
"""

print(name)

def type_text(text, delay=0.03):
    for char in text:
        print(char, end='', flush=True)
        time.sleep(delay)
    print()

text_to_type = "Loading the database....."
type_text(text_to_type)

time.sleep(0.25)

print("Database loaded")


while True:
    last_row = sheet.max_row
    box = sheet.max_row - 1
    
    Choice = questionChoicesLoop("\nWould you like too Add, Find, Remove or Open the database? (add / find / remove / open / quit):", "add", ["add", "find", "remove", "open", "quit"])

    if Choice == "add":
        print(f"The last added was box number: {box}")
        box = box + 1
        box = str(box)
        BoxNum = questionIntLoop("What is the number of this box being added:", box)
        Building = questionChoicesLoop("What building is it in (main, left, right):", "main", ["main", "left", "right"])
        Row = questionIntWithRangeLoop("which row is the item located in   (0 - 9):", "1", 0, 9)
        level = questionIntWithRangeLoop("which level is the item located on (1 - 5):", "1", 0, 5)
        Size = questionIntWithRangeLoop("What is the size of the items box  (0 - 9):", "1", 0, 9)
        Owner = questionStringLoop("Who is the owner of the box being added   :", "NAME")
        description = questionStringBlank("To not put a description press 'enter' or\nDescribe what's inside the box being added:", "")

        if Building == "main": 
            Building_num = 1
        elif Building == "left":
            Building_num = 2 
        elif Building == "right":
            Building_num = 3
        else:
            Building_num = 0

        BoxID = f"{BoxNum}{Building_num}{Row}{level}{Size}"
        BoxID = int(BoxID)

        print(f"The ID of the box is {BoxID}", sep='')

        line = BoxNum + 1

        sheet[f'A{line}'] = BoxID
        sheet[f'B{line}'] = Building
        sheet[f'C{line}'] = Row
        sheet[f'D{line}'] = level
        sheet[f'E{line}'] = Size
        sheet[f'F{line}'] = Owner

        if description == "":
            sheet[f'G{line}'] = description
        else:
            sheet[f'G{line}'] = f"Box of {description}"

        workbook.save(f"{word}.xlsx")

        print("Box added")


        border = rf"""
+------------------------------------+
| ____            _   _          _   |
|| __ )  _____  _| \ | | ___ ___| |_ |
||  _ \ / _ \ \/ |  \| |/ _ / __| __||
|| |_) | (_) >  <| |\  |  __\__ | |_ |
||____/ \___/_/\_|_| \_|\___|___/\__||
|                                    |
|           Box ID: {BoxID}           |
+------------------------------------+
"""

        label = open("label.txt", "w")
        label.write(border)
        label.close()


#        os.startfile("label.txt")
        os.startfile("label.txt", "print")
        
        print("Printing box ID sticker...")

        time.sleep(0.25)

        openDoc = str(questionString("Type 'open' to open the database or press enter to continue:","continue"))

        if openDoc == "open":
            time.sleep(0.25)
            print("Opening file...")
            os.startfile(f"{word}.xlsx")
            
            while True:
                time.sleep(1)
                close = str(questionString("Press enter to close the sheet:","close"))
                if close == "close":
                    print("closing sheet...")
                    pyautogui.hotkey('alt', 'tab')
                    pyautogui.hotkey('ctrl', 's')
                    time.sleep(0.25)
                    pyautogui.hotkey('alt', 'f4')
                    break 
        elif openDoc == "continue":
            print("Continuing...")
        else:
            print("Not opening file...")
                
    elif Choice == "find":
        print("To find a box you need to know the ID")
        find = questionChoicesLoop("Do you know the ID of the box:", "Yes", ["yes", "no"]).lower().strip()
        find = find.upper()

        if find == "YES":
            ID = questionIntLoop("What is the ID number of the box:","")
            print(f"Searching for a box with ID: {ID}...")
            time.sleep(0.50)
            for row in range(2, last_row + 1):
                cell = sheet.cell(row, 1)
                if cell.value == ID:
                    st = [1]
                    nd = [2]
                    rd = [3]
                    th = [4, 5, 6, 7, 8 ,9]

                    if sheet[f'C{row}'].value in st:
                        end = "st"
                    elif sheet[f'C{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'C{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'C{row}'].value in th:
                        end = "th"
                    
                    if sheet[f'D{row}'].value in st:
                        end = "st"
                    elif sheet[f'D{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'D{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'D{row}'].value in th:
                        end = "th"

                    print(f"Box {ID} is in row {row} of the database")
                    time.sleep(1)
                    print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                    time.sleep(3)
                    break
            else:
                print(f"A box with the ID: {ID} has not found")
                time.sleep(1)
        elif find == "NO":
            print("You need to know the ID of the box to find it")
            time.sleep(0.25)

    elif Choice == "remove":
        ID = questionIntLoop("What is the ID number of the box:","")
        print(f"Searching for a box with ID: {ID}...")
        time.sleep(0.50)
        for row in range(2, last_row + 1):
            cell = sheet.cell(row, 1)
            if cell.value == ID:
                st = [1]
                nd = [2]
                rd = [3]
                th = [4, 5, 6, 7, 8 ,9]

                if sheet[f'C{row}'].value in st:
                    end = "st"
                elif sheet[f'C{row}'].value in nd:
                    end = "nd"
                elif sheet[f'C{row}'].value in rd:
                    end = "rd"
                elif sheet[f'C{row}'].value in th:
                    end = "th"
                
                if sheet[f'D{row}'].value in st:
                    end = "st"
                elif sheet[f'D{row}'].value in nd:
                    end = "nd"
                elif sheet[f'D{row}'].value in rd:
                    end = "rd"
                elif sheet[f'D{row}'].value in th:
                    end = "th"

                print(f"Box {ID} is in row {row} of the database")
                time.sleep(1)
                print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                time.sleep(1)
                remove = questionChoicesLoop("Are you sure you want to remove this box:", "yes", ["yes", "no"]).lower().strip()
                if remove == "yes":
                    sheet.delete_rows(row, 1)
                    workbook.save(f"{word}.xlsx")
                    print("Box removed")
                    time.sleep(0.25)
                    break
                elif remove == "no":
                    print("Box not removed")
                    time.sleep(0.25)
                    break
        else:
            print(f"A box with the ID: {ID} has not found")
            time.sleep(1)

    elif Choice == "open": 
        time.sleep(0.25)
        print("Opening file...")
        os.startfile(f"{word}.xlsx")
        
        while True:
            time.sleep(1)
            close = str(questionString("Press enter to close the sheet:","close"))
            if close == "close":
                print("closing sheet...")
                pyautogui.hotkey('alt', 'tab')
                pyautogui.hotkey('ctrl', 's')
                time.sleep(0.25)
                pyautogui.hotkey('alt', 'f4')
                break 
                
    elif Choice == "quit":
        def type_text(text, delay=0.01):
            for char in text:
                print(char, end='', flush=True)
                time.sleep(delay)
            print()

        text_to_type = "Quitting program..."
        out = "Creator: Kyle Seaford"
        type_text(out)
        type_text(text_to_type)
        time.sleep(0.25)
        break
