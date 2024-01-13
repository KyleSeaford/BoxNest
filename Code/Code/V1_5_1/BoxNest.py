from openpyxl import workbook, load_workbook
import time
import os
import pyautogui
import webbrowser
import random
from datetime import datetime
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )


# User Questions

def questionString(question, defaultAnswer):
    answer = input(f"{question} [{defaultAnswer}]")
    if answer == "":
        answer = defaultAnswer
    return answer    

def questionInt(question, defaultAnswer):
    answer = questionString(question, defaultAnswer)
    answer = answer.strip()
    if answer.isdigit() == False:
        raise Exception("Invalid input, Please enter an integer")
    return int(answer)

def questionIntWithRange(question, defaultAnswer, minValue, maxValue):
    answer = questionInt(question, defaultAnswer)
    if answer > maxValue:
        raise Exception(f"Invalid input, Please enter an integer less than {maxValue}")    
    if answer <= minValue:
        raise Exception(f"Invalid input, Please enter an integer greater than {minValue}")   
    return answer

def questionChoices(question, defaultAnswer, choices):
    answer = questionString(question, defaultAnswer)
    answer = answer.lower().strip()
    if answer not in choices:
        raise Exception(f"Invalid input, Please enter one of the following: {choices}")
    return answer

def questionChoicesLoop(question, defaultAnswer, choices):
    while True:
        try:
            answer = questionChoices(question, defaultAnswer, choices)
            return answer
        except Exception as e:
            print(e)

def questionStringLoop(question, defaultAnswer):
    while True:
        try:
            answer = questionString(question, defaultAnswer)
            return answer
        except Exception as e:
            print(e)

def questionIntLoop(question, defaultAnswer):
    while True:
        try:
            answer = questionInt(question, defaultAnswer)
            return answer
        except Exception as e:
            print(e)

def questionIntWithRangeLoop(question, defaultAnswer, minValue, maxValue):
    while True:
        try:
            answer = questionIntWithRange(question, defaultAnswer, minValue, maxValue)
            return answer
        except Exception as e:
            print(e)

def questionStringBlank(question, defaultAnswer):
    answer = input(f"{question} {defaultAnswer}")
    if answer == "":
        answer = defaultAnswer
    return answer

# Load Sequence

colors = """
    0 = Black       8 = Gray
    1 = Blue        9 = Light Blue
    2 = Green       A = Light Green
    3 = Aqua        B = Light Aqua
    4 = Red         C = Light Red
    5 = Purple      D = Light Purple
    6 = Yellow      E = Light Yellow
    7 = White       F = Bright White
"""

def load_sequence():
    def color_green():
        os.system('color 0A')

    def clear():
        time.sleep(0.2)
        os.system('cls' if os.name == 'nt' else 'clear')
        time.sleep(0.1)

    def generate_loading_message():
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        random_phrase = random.choice(loading_phrases)
        return f'{current_datetime} - {random_phrase}'

    def SetUp_messages_Long():
        for _ in range(25):
            print(generate_loading_message())
            time.sleep(0.05)  

    def SetUp_messages_Short():
        for _ in range(5):
            print(generate_loading_message())
            time.sleep(0.05)  

    def generate_loading_script():
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        random_phrase = random.choice(loading_scripts)
        return f'{current_datetime} - {random_phrase}'

    def Load_Scripts():
        os.system('color 04')
        time.sleep(0.1)

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        random_phrase = random.choice(Before_loading_scripts)
        print((f'{current_datetime} - {random_phrase}'))

        for _ in loading_scripts:
            print(generate_loading_script())
            time.sleep(0.1)
        
        time.sleep(0.1)


    def setting_up():
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        random_phrase = random.choice(Before_loading_Phrases)
        print((f'{current_datetime} - {random_phrase}'))

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        random_phrase = random.choice(setUP_phrases)
        print(f'{current_datetime} - {random_phrase}')
        time.sleep(0.2)
        os.system('color 30')


    setUP_phrases = [
        "Initializing the application. Please wait while we set up the environment",
        "Loading essential components. This may take a moment as we ensure all dependencies are in place",
        "Configuring the interface layout. Getting the user interface ready for a seamless experience",
        "Loading user preferences. Applying your personalized settings to create a customized interface",
        "Retrieving UI assets. We're gathering resources to make the interface visually engaging",
        "Organizing elements. We're arranging buttons, menus, and panels for easy navigation",
        "Loading interactive widgets. Preparing dynamic components for an interactive user experience",
        "Applying theme settings. Personalizing the interface appearance according to your chosen theme",
        "Adapting to screen resolution. Ensuring a responsive design that fits your device perfectly",
        "Implementing accessibility features. Making the interface usable for all users, regardless of their abilities",
        "Fine-tuning animations. Polishing the interface with subtle and smooth visual transitions",
        "Verifying compatibility with different browsers. Ensuring consistent UI performance across platforms",
        "Optimizing layout responsiveness. Making sure the interface adapts seamlessly to different screen sizes",
        "Integrating internationalization support. Enabling the interface to be displayed in multiple languages",
        "Setting up keyboard shortcuts. Streamlining navigation through convenient key combinations",
        "Validating UI elements. Rigorously testing all user interface components for functionality and reliability",
        "Performing usability tests. Ensuring the interface is intuitive and user-friendly",
        "Creating logical information architecture. Designing the interface's structure for efficient information organization",
        "Implementing typography guidelines. Selecting appropriate fonts and styles for better readability",
        "Applying color schemes. Choosing harmonious colors that enhance the interface aesthetics and readability"
    ]

    Before_loading_scripts = [
        "ABOUT TO LOAD THE SCRIPTS...",
        "PREPARING TO INITIALIZE MODULES...",
        "CONFIGURING MODULES FOR LOADING...",
        "LOADING SCRIPTS INTO MEMORY...",
        "SETTING UP MODULES FOR LOADING...",
        "LOADING REQUIRED MODULES...",
        "CONFIGURING SCRIPT LOADING PROCEDURE...",
        "LOADING SCRIPTS AND MODULES...",
        "INITIALIZING LOADING SEQUENCE...",
        "PREPARING SCRIPTS FOR LOADING..."
    ]

    loading_scripts = [
        "Loading Main.py - Initializing the main application module",
        "Loading BoxNest.py - Configuring the nesting module for efficient box arrangement",
        "Loading userQuestions.py - Setting up the module for capturing user input and questions",
        "Loading CreateANewDatabase.py - Creating a new database structure for storing data",
        "Loading Logo.py - Initializing the module for displaying the application logo",
        "Loading Load.py - Configuring the module for displaying loading messages",
    ]

    Before_loading_Phrases = [
        "LOADING USER INTERFACE",
        "INITIALIZING UI COMPONENTS",
        "CONFIGURING UI ELEMENTS",
        "PREPARING UI FOR DISPLAY",
        "LOADING UI RESOURCES",
        "SETTING UP UI LAYOUT",
        "LOADING UI TEMPLATES",
        "CONFIGURING UI STYLES",
        "LOADING UI MODULES",
        "INITIALIZING UI CONTROLS"
    ]

    loading_phrases = [
        "Initializing the application. Please wait while we set up the environment",
        "Loading essential components. This may take a moment as we ensure all dependencies are in place.",
        "Preparing the system for optimal performance. Hold on as we configure various settings",
        "Compiling necessary code modules. This process might take a while, so make yourself comfortable",
        "Optimizing resource allocation. We are carefully managing system resources for improved efficiency",
        "Verifying user permissions. Sit tight while we authenticate your access rights",
        "Establishing secure connections. We're ensuring a safe connection to our servers",
        "Loading the user interface. Relax as we create an intuitive and user-friendly experience.",
        "Retrieving updated data. We're fetching the latest information for you",
        "Performing data integrity checks. Your data is important to us, so we're ensuring its accuracy"
    ]


    color_green()
    SetUp_messages_Long()
    Load_Scripts()
    SetUp_messages_Short()
    setting_up()
    clear()


def clear_cmd():
    time.sleep(0.3)
    os.system('cls' if os.name == 'nt' else 'clear')
    time.sleep(0.2)

# Logos

def Logo():
    name = f"""
             88888"88b                    8888b   888                   888    
             888  .88P                    88888b  888                   888          
             8888888K.   .d88b.  888  888 888Y88b 888  .d88b.  .d8888b  888888       Organized to Perfection
             888  "Y88b d88""88b `Y8bd8P' 888 Y88b888 d8P  Y8b 88K      888          >>===================<<
             888    888 888  888   X88K   888  Y88888 88888888 "Y8888b. 888             By: Kyle Seaford
             888   d88P Y88..88P .d8""8b. 888   Y8888 Y8b.          X88 Y88b.           
             8888888P"   "Y88P"  888  888 888    Y888  "Y8888   88888P'  "Y888 
    """

    print(name)


def Choice_Box():
    Choice1 = r"""
            /=======================================Welcome to BoxNest======================================\      
           ||                                                                      ########                 || 
           ||                                                                  #######  ######              || 
           ||      Options:                                                    #####     ##########         || 
           ||      1. "Add" To The Database [Enter]                       ######    ##################      || 
           ||      2. "Find" In The Database                              #####  ###########     #####      ||         
           ||      3. "Remove" From The Database                          ## ############     ##### ##      || 
           ||      4. "Open" The Database                                 ##   #######    ######    ##      || 
           ||      5. "Quit" The Program                                  ##   ##############       ##      || 
           ||                                                             ##   #####   ##        ## ##      || 
           ||      Other:                                                 ##      #    ##        ## ##      || 
           ||      6. Open The "Website"                                   ######      ##       ######      || 
           ||      7. "Print" The Database                                     #####   ##    #####          || 
           ||                                                                    ##############             || 
           ||                                                                        ########               ||
            \===============================================================================================/
    """
    print(Choice1)


# Main

word = "database"

workbook = load_workbook(f"{word}.xlsx")
sheet = workbook.active

clear_cmd()
load_sequence()
clear_cmd()
Logo()

while True:
    last_row = sheet.max_row
    box = sheet.max_row - 1

    Choice_Box()
    
    Choice = questionChoicesLoop("Please Select An Option:", "add", ["add", "find", "remove", "open", "quit", "website", "print", "1", "2", "3", "4", "5", "6", "7"]).lower().strip()

    if Choice == "add" or Choice == "1":
        print(f"The last added was box number: {box}")
        box = box + 1
        box = str(box)
        BoxNum = questionIntLoop("What is the number of this box being added:", box)
        Building = questionChoicesLoop("What building is it in (main, left, right):", "main", ["main", "left", "right"])
        Row = questionIntWithRangeLoop("which row is the item located in   (0 - 9):", "1", 0, 9)
        level = questionIntWithRangeLoop("which level is the item located on (1 - 5):", "1", 0, 5)
        Size = questionIntWithRangeLoop("What is the size of the items box  (0 - 9):", "1", 0, 9)
        Owner = questionStringLoop("Who is the owner of the box being added   :", "NAME")
        description = questionStringBlank("To not put a description press 'enter' or\nDescribe what's inside the box being added:", "")

        if Building == "main": 
            Building_num = 1
        elif Building == "left":
            Building_num = 2 
        elif Building == "right":
            Building_num = 3
        else:
            Building_num = 0

        BoxID = f"{BoxNum}{Building_num}{Row}{level}{Size}"
        BoxID = int(BoxID)

        print(f"The ID of the box is {BoxID}", sep='')

        line = BoxNum + 1

        sheet[f'A{line}'] = BoxID
        sheet[f'B{line}'] = Building
        sheet[f'C{line}'] = Row
        sheet[f'D{line}'] = level
        sheet[f'E{line}'] = Size
        sheet[f'F{line}'] = Owner

        if description == "":
            sheet[f'G{line}'] = description
        else:
            sheet[f'G{line}'] = f"Box of {description}"

        workbook.save(f"{word}.xlsx")

        print("Box added")


        border = rf"""
+------------------------------------+
| ____            _   _          _   |
|| __ )  _____  _| \ | | ___ ___| |_ |
||  _ \ / _ \ \/ |  \| |/ _ / __| __||
|| |_) | (_) >  <| |\  |  __\__ | |_ |
||____/ \___/_/\_|_| \_|\___|___/\__||
|                                    |
|           Box ID: {BoxID}           |
+------------------------------------+
"""

        label = open("label.txt", "w")
        label.write(border)
        label.close()


#        os.startfile("label.txt")
        os.startfile("label.txt", "print")
        
        print("Printing box ID sticker...")

        time.sleep(0.25)

        openDoc = str(questionString("Type 'open' to open the database or press enter to continue:","continue"))

        if openDoc == "open":
            time.sleep(0.25)
            print("Opening file...")
            os.startfile(f"{word}.xlsx")
            
            while True:
                time.sleep(1)
                close = str(questionString("Press enter to close the sheet:","close"))
                if close == "close":
                    print("closing sheet...")
                    pyautogui.hotkey('alt', 'tab')
                    pyautogui.hotkey('ctrl', 's')
                    time.sleep(0.25)
                    pyautogui.hotkey('alt', 'f4')
                    clear_cmd()
                    break 
        elif openDoc == "continue":
            print("Continuing...")
            clear_cmd()
        else:
            print("Not opening file...")
            clear_cmd()
                
    elif Choice == "find" or Choice == "2":
        print("To find a box you need to know the ID")
        find = questionChoicesLoop("Do you know the ID of the box:", "Yes", ["yes", "no"]).lower().strip()
        find = find.upper()

        if find == "YES":
            ID = questionIntLoop("What is the ID number of the box:","")
            print(f"Searching for a box with ID: {ID}...")
            time.sleep(0.50)
            for row in range(2, last_row + 1):
                cell = sheet.cell(row, 1)
                if cell.value == ID:
                    st = [1]
                    nd = [2]
                    rd = [3]
                    th = [4, 5, 6, 7, 8 ,9]

                    if sheet[f'C{row}'].value in st:
                        end = "st"
                    elif sheet[f'C{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'C{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'C{row}'].value in th:
                        end = "th"
                    
                    if sheet[f'D{row}'].value in st:
                        end = "st"
                    elif sheet[f'D{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'D{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'D{row}'].value in th:
                        end = "th"

                    print(f"Box {ID} is in row {row} of the database")
                    time.sleep(1)
                    print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                    time.sleep(3)
                    clear_cmd()
                    break
            else:
                print(f"A box with the ID: {ID} has not found")
                time.sleep(1)
                clear_cmd()
        elif find == "NO":
            print("You need to know the ID of the box to find it")
            time.sleep(0.5)
            clear_cmd()

    elif Choice == "remove" or Choice == "3":
        ID = questionIntLoop("What is the ID number of the box:","")
        print(f"Searching for a box with ID: {ID}...")
        time.sleep(0.50)
        for row in range(2, last_row + 1):
            cell = sheet.cell(row, 1)
            if cell.value == ID:
                st = [1]
                nd = [2]
                rd = [3]
                th = [4, 5, 6, 7, 8 ,9]

                if sheet[f'C{row}'].value in st:
                    end = "st"
                elif sheet[f'C{row}'].value in nd:
                    end = "nd"
                elif sheet[f'C{row}'].value in rd:
                    end = "rd"
                elif sheet[f'C{row}'].value in th:
                    end = "th"
                
                if sheet[f'D{row}'].value in st:
                    end = "st"
                elif sheet[f'D{row}'].value in nd:
                    end = "nd"
                elif sheet[f'D{row}'].value in rd:
                    end = "rd"
                elif sheet[f'D{row}'].value in th:
                    end = "th"

                print(f"Box {ID} is in row {row} of the database")
                time.sleep(1)
                print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                time.sleep(1)
                remove = questionChoicesLoop("Are you sure you want to remove this box:", "yes", ["yes", "no"]).lower().strip()
                if remove == "yes":
                    sheet.delete_rows(row, 1)
                    workbook.save(f"{word}.xlsx")
                    print("Box removed")
                    time.sleep(0.5)
                    clear_cmd()
                    break
                elif remove == "no":
                    print("Box not removed")
                    time.sleep(0.5)
                    clear_cmd()
                    break
        else:
            print(f"A box with the ID: {ID} has not found")
            time.sleep(1)
            clear_cmd()

    elif Choice == "open" or Choice == "4": 
        time.sleep(0.25)
        print("Opening file...")
        os.startfile(f"{word}.xlsx")
        
        while True:
            time.sleep(1)
            close = str(questionString("Press enter to close the sheet:","close"))
            if close == "close":
                print("closing sheet...")
                pyautogui.hotkey('alt', 'tab')
                pyautogui.hotkey('ctrl', 's')
                time.sleep(0.25)
                pyautogui.hotkey('alt', 'f4')
                time.sleep(0.25)
                clear_cmd()
                break 
                
    elif Choice == "quit" or Choice == "5":
        def type_text(text, delay=0.01):
            for char in text:
                print(char, end='', flush=True)
                time.sleep(delay)
            print()

        Quitting = "Quitting program..."
        Thanks = "Thank you for using BoxNest"
        Creator = "Creator: Kyle Seaford"
        type_text(Thanks)
        type_text(Creator)
        type_text(Quitting)
        time.sleep(0.25)
        break


    elif Choice == "website" or Choice == "6":
        print("Opening website...")
        webbrowser.open('http://test.kyle-seaford.co.uk/index.html')
        time.sleep(0.5)
        clear_cmd()

    elif Choice == "print" or Choice == "7":
        print("printing database...")
#        os.startfile("database.xlsx")
        os.startfile("database.xlsx", "print")
        time.sleep(0.5)
        clear_cmd()
