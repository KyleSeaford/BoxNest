import shutil
import os
import time
import logging
from openpyxl import load_workbook

startTime = time.time()

logging.basicConfig(level=logging.DEBUG,  
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='backup+log/BoxNest.log', filemode='w')


main_db_path = 'database.xlsx'
backup_folder = 'backup+log/'

# Backup database
def backup_database():
    backup_filename = f'{time.strftime("%Y-%m-%d_%H-%M")}_databaseBackup.xlsx'
    backup_path = os.path.join(backup_folder, backup_filename)

    shutil.copy(main_db_path, backup_path)


# Backup database if the program has been running for 2 hours
def should_backup_database(current_time, max_runtime_seconds):
    elapsed_time = current_time - startTime

    return elapsed_time >= max_runtime_seconds


# Backup database if the Database has more than 30 rows
def should_backup_database_rows(min_rows_for_backup):
    last_row = get_last_row()

    return last_row >= min_rows_for_backup

def get_last_row():
    workbook = load_workbook(main_db_path)
    sheet = workbook.active

    last_row = sheet.max_row

    workbook.close()

    return last_row


# Backup database if the program has been running for 2 hours or if the Database has more than 30 rows
def CheckBackup():
    if should_backup_database(time.time(), 2 * 60 * 60):
        logging.info("Time-based condition triggered: Backup needed, backed up database")
        backup_database()
    elif should_backup_database_rows(30):
        logging.info("Row-based condition triggered: Backup needed, backed up database")
        backup_database()
    else:
        logging.info("No conditions triggered: Database backup not needed")
    
